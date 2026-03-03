#!/usr/bin/env python3
"""
Extract ALL condo-related rows from BIR zonal value NCR workbooks.
Scans for classification codes RC, CC, PS, PH, PC and condo keyword indicators.
Outputs raw cell data with merged cell context.
"""

import os
import re
import sys
from pathlib import Path

import xlrd
import openpyxl

BASE_DIR = Path("/home/runner/work/me/me/loops/reverse-ph-zonal-value-engine/input/bir-workbook-samples/extracted")

TARGET_FILES = [
    "RDO No. 30 - Binondo.xls",
    "RDO No. 41 - Mandaluyong City.xls",
    "RDO No. 44 - Taguig-Pateros.xlsx",
    "RDO No. 47 - East Makati.xls",
    "RDO No. 48 - West Makati.xls",
    "RDO No. 49 - North Makati City.xls",
    "RDO No. 50 - South Makati.xls",
    "RDO No. 51 - Pasay City.xls",
    "RDO No. 52 - Paranaque City.xls",
    "RDO No. 43 - Pasig City.xls",
]

# Classification codes that indicate condos
CONDO_CODES = {"RC", "CC", "PS", "PH", "PC"}

# Keywords in street/name columns that indicate condos
CONDO_KEYWORDS = re.compile(
    r"CONDO|CONDOMINIUM|TOWER|BUILDING|BLDG\.?|PENTHOUSE|RESIDENCES|"
    r"UNIT|PARKING|FAR\s*[-=]|SQM|PER\s+SQ|FLOOR\s+AREA|"
    r"BASEMENT|PODIUM",
    re.IGNORECASE,
)

CONTEXT_ROWS = 3  # rows before/after a condo hit


def cell_value_str(val):
    """Convert cell value to string, handling floats that are ints."""
    if val is None or val == "":
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def read_xls_sheet(filepath, sheet_index=1):
    """Read an .xls file sheet, return rows as list of lists + merged cells."""
    wb = xlrd.open_workbook(str(filepath), formatting_info=True)
    if wb.nsheets <= sheet_index:
        print(f"  WARNING: {filepath.name} has only {wb.nsheets} sheets, using sheet 0")
        sheet_index = 0
    sh = wb.sheet_by_index(sheet_index)
    print(f"  Sheet name: '{sh.name}', rows={sh.nrows}, cols={sh.ncols}")

    rows = []
    for r in range(sh.nrows):
        row = []
        for c in range(sh.ncols):
            cell = sh.cell(r, c)
            row.append(cell_value_str(cell.value))
        rows.append(row)

    # Merged cells: list of (rlo, rhi, clo, chi)
    merged = sh.merged_cells  # list of (row_lo, row_hi, col_lo, col_hi)
    return rows, list(merged), sh.name


def read_xlsx_sheet(filepath, sheet_index=1):
    """Read an .xlsx file sheet, return rows as list of lists + merged cells."""
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    sheet_names = wb.sheetnames
    if len(sheet_names) <= sheet_index:
        print(f"  WARNING: {filepath.name} has only {len(sheet_names)} sheets, using sheet 0")
        sheet_index = 0
    sh = wb[sheet_names[sheet_index]]
    print(f"  Sheet name: '{sh.title}', rows={sh.max_row}, cols={sh.max_column}")

    rows = []
    for row in sh.iter_rows(min_row=1, max_row=sh.max_row, max_col=sh.max_column):
        rows.append([cell_value_str(cell.value) for cell in row])

    # Merged cells
    merged = []
    for mc in sh.merged_cells.ranges:
        merged.append((mc.min_row - 1, mc.max_row, mc.min_col - 1, mc.max_col))  # convert to 0-indexed like xlrd
    return rows, merged, sh.title


def is_condo_row(row):
    """Check if a row contains condo indicators. Returns list of reasons."""
    reasons = []
    row_text = " ".join(row).upper()

    # Check each cell for classification codes
    for i, cell in enumerate(row):
        val = cell.strip().upper()
        if val in CONDO_CODES:
            reasons.append(f"col{i}='{val}'(condo_code)")

    # Check for keyword matches in the full row text
    matches = CONDO_KEYWORDS.findall(row_text)
    if matches:
        unique_matches = list(set(m.upper() for m in matches))
        reasons.append(f"keywords={unique_matches}")

    return reasons


def find_merged_cells_for_range(merged_cells, row_start, row_end):
    """Find all merged cell ranges that overlap with a given row range."""
    overlapping = []
    for mc in merged_cells:
        rlo, rhi, clo, chi = mc
        if rlo < row_end and rhi > row_start:
            overlapping.append(mc)
    return overlapping


def format_merged_cell(mc):
    """Format a merged cell range as a readable string."""
    rlo, rhi, clo, chi = mc
    return f"rows[{rlo}:{rhi}] cols[{clo}:{chi}] ({rhi-rlo}r x {chi-clo}c)"


def format_row(row, row_idx, max_cols=None):
    """Format a row for display, showing column indices."""
    parts = []
    for i, val in enumerate(row):
        if val:
            parts.append(f"[{i}]={val}")
    if not parts:
        return f"  Row {row_idx}: (empty)"
    return f"  Row {row_idx}: " + " | ".join(parts)


def extract_condo_data(filepath):
    """Extract all condo-related rows from a workbook."""
    ext = filepath.suffix.lower()
    if ext == ".xls":
        rows, merged, sheet_name = read_xls_sheet(filepath)
    elif ext == ".xlsx":
        rows, merged, sheet_name = read_xlsx_sheet(filepath)
    else:
        print(f"  SKIP: Unknown extension {ext}")
        return

    num_rows = len(rows)
    num_cols = max(len(r) for r in rows) if rows else 0

    # First pass: find the header row to understand column layout
    print(f"\n  --- HEADER DETECTION ---")
    for i, row in enumerate(rows[:20]):
        row_text = " ".join(row).upper()
        if any(kw in row_text for kw in ["ZONE", "STREET", "CLASS", "BARANGAY", "VALUE", "MARKET"]):
            print(f"  Possible header at row {i}: {format_row(row, i)}")

    # Second pass: find all condo rows
    condo_row_indices = set()
    condo_reasons = {}

    for i, row in enumerate(rows):
        reasons = is_condo_row(row)
        if reasons:
            condo_row_indices.add(i)
            condo_reasons[i] = reasons

    if not condo_row_indices:
        print(f"  NO CONDO ROWS FOUND")
        return

    # Expand to include context rows
    expanded_indices = set()
    for idx in sorted(condo_row_indices):
        for offset in range(-CONTEXT_ROWS, CONTEXT_ROWS + 1):
            r = idx + offset
            if 0 <= r < num_rows:
                expanded_indices.add(r)

    # Build contiguous blocks
    sorted_indices = sorted(expanded_indices)
    blocks = []
    current_block = [sorted_indices[0]]
    for idx in sorted_indices[1:]:
        if idx == current_block[-1] + 1:
            current_block.append(idx)
        else:
            blocks.append(current_block)
            current_block = [idx]
    blocks.append(current_block)

    print(f"\n  --- CONDO ROWS FOUND: {len(condo_row_indices)} matches in {len(blocks)} blocks ---")

    # Output all merged cells in the sheet that are relevant
    all_condo_merged = find_merged_cells_for_range(
        merged, min(expanded_indices), max(expanded_indices) + 1
    )

    if all_condo_merged:
        print(f"\n  --- MERGED CELLS OVERLAPPING CONDO BLOCKS ({len(all_condo_merged)} total) ---")
        for mc in sorted(all_condo_merged):
            rlo, rhi, clo, chi = mc
            # Show the value in the top-left cell
            if rlo < len(rows) and clo < len(rows[rlo]):
                val = rows[rlo][clo]
            else:
                val = ""
            print(f"    {format_merged_cell(mc)}  value='{val}'")

    # Output each block
    for block_num, block in enumerate(blocks):
        print(f"\n  === BLOCK {block_num + 1} (rows {block[0]}-{block[-1]}) ===")

        # Show merged cells specific to this block
        block_merged = find_merged_cells_for_range(merged, block[0], block[-1] + 1)
        if block_merged:
            print(f"  Merged cells in block:")
            for mc in sorted(block_merged):
                rlo, rhi, clo, chi = mc
                if rlo < len(rows) and clo < len(rows[rlo]):
                    val = rows[rlo][clo]
                else:
                    val = ""
                print(f"    {format_merged_cell(mc)}  value='{val}'")

        for idx in block:
            marker = ""
            if idx in condo_reasons:
                marker = f"  <<<< MATCH: {condo_reasons[idx]}"
            print(f"{format_row(rows[idx], idx)}{marker}")

    # Additional: dump ALL merged cells in the sheet for understanding structure
    print(f"\n  --- ALL MERGED CELLS IN SHEET ({len(merged)} total) ---")
    # Group by row range size to find patterns
    merge_patterns = {}
    for mc in sorted(merged):
        rlo, rhi, clo, chi = mc
        key = f"{rhi-rlo}r x {chi-clo}c"
        if key not in merge_patterns:
            merge_patterns[key] = []
        merge_patterns[key].append(mc)

    for pattern, cells in sorted(merge_patterns.items()):
        print(f"  Pattern {pattern}: {len(cells)} occurrences")
        # Show first 5 examples
        for mc in cells[:5]:
            rlo, rhi, clo, chi = mc
            if rlo < len(rows) and clo < len(rows[rlo]):
                val = rows[rlo][clo][:80] if rows[rlo][clo] else ""
            else:
                val = ""
            print(f"    rows[{rlo}:{rhi}] cols[{clo}:{chi}] = '{val}'")
        if len(cells) > 5:
            print(f"    ... and {len(cells) - 5} more")

    return condo_row_indices


def main():
    print("=" * 120)
    print("BIR ZONAL VALUE WORKBOOK - CONDO DATA EXTRACTION")
    print("=" * 120)

    for filename in TARGET_FILES:
        filepath = BASE_DIR / filename
        if not filepath.exists():
            print(f"\nERROR: File not found: {filepath}")
            continue

        print(f"\n{'#' * 120}")
        print(f"# FILE: {filename}")
        print(f"# Path: {filepath}")
        print(f"{'#' * 120}")

        try:
            extract_condo_data(filepath)
        except Exception as e:
            print(f"  ERROR processing {filename}: {e}")
            import traceback
            traceback.print_exc()

    print(f"\n{'=' * 120}")
    print("EXTRACTION COMPLETE")
    print(f"{'=' * 120}")


if __name__ == "__main__":
    main()
