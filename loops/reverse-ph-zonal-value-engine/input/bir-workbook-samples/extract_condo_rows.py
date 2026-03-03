#!/usr/bin/env python3
"""
Extract ALL condo-related rows from provincial BIR zonal value workbooks.
For each workbook, reads the SECOND sheet (index 1).
Outputs raw cell data with surrounding context rows.
"""

import os
import re
import json
from pathlib import Path

import xlrd
import openpyxl

BASE_DIR = Path("/home/runner/work/me/me/loops/reverse-ph-zonal-value-engine/input/bir-workbook-samples/extracted-provincial")

TARGET_FILES = [
    "RDO No. 81 - Cebu City North.xlsx",
    "RDO No. 83 -Talisay City, CebuWeb.xls",
    "RDO No. 56 - Calamba City, Central Laguna.xls",
    "RDO NO. 113A - West Davao City.xlsx",
    "RDO No. 4 - Calasiao, Central Pangasinan.xls",
    "RDO No. 5 - Alaminos City, West Pangasinan.xls",
]

# Classification codes that indicate condos
CONDO_CODES = {"RC", "CC", "PS", "PH", "PC"}

# Text patterns in street/location/name columns
CONDO_TEXT_PATTERNS = re.compile(
    r"CONDO|CONDOMINIUM|TOWER|BUILDING|TOWNHOUSE|TOWN\s*HOUSE|CCT|PARKING\s*SLOT",
    re.IGNORECASE,
)

# Special patterns to watch for
SPECIAL_PATTERNS = re.compile(
    r"ALL\s*OTHER\s*CONDO|HOLLAND\s*PARK|MY\s*CUBE|NEWLY\s*CREATED",
    re.IGNORECASE,
)

CONTEXT_ROWS = 3  # rows before and after each condo hit


def cell_str(val):
    """Convert cell value to string for display, preserving type info."""
    if val is None:
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val).strip()


def is_condo_row(row_values):
    """Check if any cell in the row indicates condo content."""
    reasons = []
    for i, val in enumerate(row_values):
        s = cell_str(val).upper().strip()
        if not s:
            continue
        # Check classification codes (exact match on short values)
        if s in CONDO_CODES:
            reasons.append(f"col[{i}] code='{s}'")
        # Check text patterns
        if CONDO_TEXT_PATTERNS.search(s):
            reasons.append(f"col[{i}] text match in '{cell_str(val)[:80]}'")
    return reasons


def read_xlsx(filepath):
    """Read .xlsx file using openpyxl. Returns sheet info dict."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_names = wb.sheetnames
    print(f"  Sheet names: {sheet_names}")

    if len(sheet_names) < 2:
        print(f"  WARNING: Only {len(sheet_names)} sheet(s). Using index 0.")
        ws = wb[sheet_names[0]]
    else:
        ws = wb[sheet_names[1]]
        print(f"  Using sheet: '{sheet_names[1]}' (index 1)")

    # Gather merged cell ranges
    merged_ranges = []
    for mc in ws.merged_cells.ranges:
        merged_ranges.append({
            "range": str(mc),
            "min_row": mc.min_row,
            "max_row": mc.max_row,
            "min_col": mc.min_col,
            "max_col": mc.max_col,
        })

    # Read all rows
    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
        row_data = []
        for cell in row:
            row_data.append({
                "value": cell.value,
                "coordinate": cell.coordinate,
            })
        all_rows.append(row_data)

    # Also read header area (first 15 rows) for preamble/footnotes
    header_rows = []
    for i, row in enumerate(all_rows[:15]):
        vals = [cell_str(c["value"]) for c in row]
        if any(v for v in vals):
            header_rows.append({"row_idx": i, "values": vals})

    # Read tail area (last 30 rows) for footnotes
    tail_rows = []
    for i in range(max(0, len(all_rows) - 30), len(all_rows)):
        row = all_rows[i]
        vals = [cell_str(c["value"]) for c in row]
        if any(v for v in vals):
            tail_rows.append({"row_idx": i, "values": vals})

    wb.close()
    return {
        "all_rows": all_rows,
        "merged_ranges": merged_ranges,
        "header_rows": header_rows,
        "tail_rows": tail_rows,
        "max_col": ws.max_column,
        "max_row": ws.max_row,
    }


def read_xls(filepath):
    """Read .xls file using xlrd. Returns sheet info dict."""
    wb = xlrd.open_workbook(filepath, formatting_info=True)
    sheet_names = wb.sheet_names()
    print(f"  Sheet names: {sheet_names}")

    if len(sheet_names) < 2:
        print(f"  WARNING: Only {len(sheet_names)} sheet(s). Using index 0.")
        ws = wb.sheet_by_index(0)
    else:
        ws = wb.sheet_by_index(1)
        print(f"  Using sheet: '{sheet_names[1]}' (index 1)")

    # Gather merged cell ranges
    merged_ranges = []
    for (rlo, rhi, clo, chi) in ws.merged_cells:
        merged_ranges.append({
            "range": f"R{rlo+1}C{clo+1}:R{rhi}C{chi}",
            "min_row": rlo + 1,
            "max_row": rhi,
            "min_col": clo + 1,
            "max_col": chi,
        })

    # Read all rows
    all_rows = []
    for rx in range(ws.nrows):
        row_data = []
        for cx in range(ws.ncols):
            cell = ws.cell(rx, cx)
            val = cell.value
            # xlrd returns floats for numbers
            row_data.append({
                "value": val,
                "coordinate": f"R{rx+1}C{cx+1}",
            })
        all_rows.append(row_data)

    # Header area
    header_rows = []
    for i in range(min(15, len(all_rows))):
        row = all_rows[i]
        vals = [cell_str(c["value"]) for c in row]
        if any(v for v in vals):
            header_rows.append({"row_idx": i, "values": vals})

    # Tail area
    tail_rows = []
    for i in range(max(0, len(all_rows) - 30), len(all_rows)):
        row = all_rows[i]
        vals = [cell_str(c["value"]) for c in row]
        if any(v for v in vals):
            tail_rows.append({"row_idx": i, "values": vals})

    wb.release_resources()
    return {
        "all_rows": all_rows,
        "merged_ranges": merged_ranges,
        "header_rows": header_rows,
        "tail_rows": tail_rows,
        "max_col": ws.ncols,
        "max_row": ws.nrows,
    }


def find_condo_rows(sheet_data):
    """Find all condo-related rows and their context."""
    all_rows = sheet_data["all_rows"]
    total_rows = len(all_rows)

    # First pass: find all condo row indices
    condo_hits = {}
    for idx, row in enumerate(all_rows):
        values = [c["value"] for c in row]
        reasons = is_condo_row(values)
        if reasons:
            condo_hits[idx] = reasons

    # Also scan for special patterns anywhere
    special_hits = {}
    for idx, row in enumerate(all_rows):
        for c in row:
            s = cell_str(c["value"])
            if s and SPECIAL_PATTERNS.search(s):
                if idx not in special_hits:
                    special_hits[idx] = []
                special_hits[idx].append(f"SPECIAL: '{s[:100]}'")

    # Merge special hits into condo hits
    for idx, reasons in special_hits.items():
        if idx in condo_hits:
            condo_hits[idx].extend(reasons)
        else:
            condo_hits[idx] = reasons

    # Build context ranges (merge overlapping)
    context_indices = set()
    for idx in condo_hits:
        for i in range(max(0, idx - CONTEXT_ROWS), min(total_rows, idx + CONTEXT_ROWS + 1)):
            context_indices.add(i)

    # Group into contiguous blocks
    sorted_indices = sorted(context_indices)
    blocks = []
    if sorted_indices:
        current_block = [sorted_indices[0]]
        for i in range(1, len(sorted_indices)):
            if sorted_indices[i] == sorted_indices[i - 1] + 1:
                current_block.append(sorted_indices[i])
            else:
                blocks.append(current_block)
                current_block = [sorted_indices[i]]
        blocks.append(current_block)

    # Find merged cells that overlap with condo rows
    relevant_merges = []
    for mc in sheet_data["merged_ranges"]:
        for idx in condo_hits:
            # Check if merge overlaps with any condo row (0-indexed vs 1-indexed)
            row_1indexed = idx + 1
            if mc["min_row"] <= row_1indexed <= mc["max_row"]:
                relevant_merges.append(mc)
                break
            # Also check context
            for ci in range(max(0, idx - CONTEXT_ROWS), min(total_rows, idx + CONTEXT_ROWS + 1)):
                ci_1indexed = ci + 1
                if mc["min_row"] <= ci_1indexed <= mc["max_row"]:
                    relevant_merges.append(mc)
                    break

    # Scan for preamble/footnote text about condos
    condo_notes = []
    for idx, row in enumerate(all_rows):
        for c in row:
            s = cell_str(c["value"]).upper()
            if any(kw in s for kw in ["CCT", "GROUND FLOOR", "RC TO CC", "RC→CC", "CONDOMINIUM CERTIFICATE",
                                       "PARKING", "COMMON AREA", "PREAMBLE", "NOTE:", "***", "****",
                                       "NEWLY CREATED", "UPGRADE"]):
                if len(cell_str(c["value"])) > 5:  # skip short matches
                    condo_notes.append({
                        "row_idx": idx,
                        "coordinate": c["coordinate"],
                        "text": cell_str(c["value"])[:500],
                    })

    return {
        "condo_hits": condo_hits,
        "blocks": blocks,
        "relevant_merges": relevant_merges,
        "condo_notes": condo_notes,
    }


def format_row(row_data, row_idx, is_hit=False, hit_reasons=None):
    """Format a single row for display."""
    values = [cell_str(c["value"]) for c in row_data]
    marker = ">>>" if is_hit else "   "
    row_num = row_idx + 1  # 1-indexed for display

    # Build compact representation
    cells = []
    for i, v in enumerate(values):
        if v:
            cells.append(f"[{i}]={v}")

    line = f"  {marker} Row {row_num:4d}: {' | '.join(cells)}"
    if hit_reasons:
        line += f"\n           MATCH: {'; '.join(hit_reasons)}"
    return line


def process_workbook(filename):
    """Process a single workbook and return results."""
    filepath = BASE_DIR / filename
    print(f"\n{'='*120}")
    print(f"FILE: {filename}")
    print(f"{'='*120}")

    if not filepath.exists():
        print(f"  FILE NOT FOUND: {filepath}")
        return

    ext = filepath.suffix.lower()
    if ext == ".xlsx":
        sheet_data = read_xlsx(filepath)
    elif ext == ".xls":
        sheet_data = read_xls(filepath)
    else:
        print(f"  Unknown extension: {ext}")
        return

    print(f"  Dimensions: {sheet_data['max_row']} rows x {sheet_data['max_col']} cols")
    print(f"  Total merged ranges: {len(sheet_data['merged_ranges'])}")

    # Show header
    print(f"\n  --- HEADER AREA (first 15 rows) ---")
    for h in sheet_data["header_rows"]:
        vals = [f"[{i}]={v}" for i, v in enumerate(h["values"]) if v]
        print(f"    Row {h['row_idx']+1}: {' | '.join(vals)}")

    # Find condo rows
    results = find_condo_rows(sheet_data)

    print(f"\n  --- CONDO HITS: {len(results['condo_hits'])} rows ---")

    if not results["condo_hits"]:
        print("    No condo-related rows found.")
        # Still show tail for footnotes
        print(f"\n  --- TAIL AREA (last 30 rows) ---")
        for t in sheet_data["tail_rows"]:
            vals = [f"[{i}]={v}" for i, v in enumerate(t["values"]) if v]
            if vals:
                print(f"    Row {t['row_idx']+1}: {' | '.join(vals)}")
        return

    # Show blocks with context
    print(f"\n  --- CONDO BLOCKS ({len(results['blocks'])} contiguous groups) ---")
    for bi, block in enumerate(results["blocks"]):
        print(f"\n  === Block {bi+1} (rows {block[0]+1}-{block[-1]+1}) ===")
        for idx in block:
            row = sheet_data["all_rows"][idx]
            is_hit = idx in results["condo_hits"]
            reasons = results["condo_hits"].get(idx)
            print(format_row(row, idx, is_hit, reasons))

    # Show relevant merged cells
    if results["relevant_merges"]:
        print(f"\n  --- MERGED CELLS near condo rows ({len(results['relevant_merges'])}) ---")
        for mc in results["relevant_merges"]:
            print(f"    {mc['range']} (rows {mc['min_row']}-{mc['max_row']}, cols {mc['min_col']}-{mc['max_col']})")

    # Show condo-related notes/preambles
    if results["condo_notes"]:
        print(f"\n  --- CONDO-RELATED NOTES/PREAMBLE ({len(results['condo_notes'])}) ---")
        seen = set()
        for note in results["condo_notes"]:
            key = note["text"][:100]
            if key not in seen:
                seen.add(key)
                print(f"    Row {note['row_idx']+1} ({note['coordinate']}): {note['text']}")

    # Show tail area for footnotes
    print(f"\n  --- TAIL AREA (last 30 rows with content) ---")
    for t in sheet_data["tail_rows"]:
        vals = [f"[{i}]={v}" for i, v in enumerate(t["values"]) if v]
        if vals:
            print(f"    Row {t['row_idx']+1}: {' | '.join(vals)}")

    # Summary stats
    print(f"\n  --- SUMMARY ---")
    print(f"    Total condo-related rows: {len(results['condo_hits'])}")
    code_counts = {}
    for idx, reasons in results["condo_hits"].items():
        for r in reasons:
            if "code=" in r:
                code = r.split("code='")[1].rstrip("'")
                code_counts[code] = code_counts.get(code, 0) + 1
    if code_counts:
        print(f"    Classification code counts: {code_counts}")

    # Detailed dump of each condo hit row (full cell data)
    print(f"\n  --- DETAILED CELL DATA FOR EACH CONDO HIT ---")
    for idx in sorted(results["condo_hits"].keys()):
        row = sheet_data["all_rows"][idx]
        reasons = results["condo_hits"][idx]
        print(f"\n    Row {idx+1} (match: {'; '.join(reasons)}):")
        for ci, cell in enumerate(row):
            v = cell_str(cell["value"])
            if v:
                print(f"      Col {ci:2d} ({cell['coordinate']}): '{v}'")


def main():
    print("=" * 120)
    print("BIR PROVINCIAL WORKBOOK - CONDO ROW EXTRACTION")
    print("=" * 120)

    for filename in TARGET_FILES:
        try:
            process_workbook(filename)
        except Exception as e:
            print(f"\n  ERROR processing {filename}: {e}")
            import traceback
            traceback.print_exc()

    print("\n" + "=" * 120)
    print("EXTRACTION COMPLETE")
    print("=" * 120)


if __name__ == "__main__":
    main()
