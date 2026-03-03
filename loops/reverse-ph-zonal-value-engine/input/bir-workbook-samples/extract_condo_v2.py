#!/usr/bin/env python3
"""
Extract ALL condo-related rows from BIR zonal value NCR workbooks.
V2: More structured output, focused on actual condo entries with their data rows.
Produces a single structured output file per workbook section.
"""

import os
import re
import sys
from pathlib import Path
from collections import defaultdict

import xlrd
import openpyxl

BASE_DIR = Path("/home/runner/work/me/me/loops/reverse-ph-zonal-value-engine/input/bir-workbook-samples/extracted")
OUT_DIR = Path("/home/runner/work/me/me/loops/reverse-ph-zonal-value-engine/input/bir-workbook-samples")

TARGET_FILES = [
    "RDO No. 30 - Binondo.xls",
    "RDO No. 41 - Mandaluyong City.xls",
    "RDO No. 44 - Taguig-Pateros.xlsx",
    "RDO No. 47 - East Makati.xls",
    "RDO No. 48 - West Makati.xls",
    "RDO No. 49 - North Makati City.xls",
    "RDO No. 50 - South Makati.xls",
    "RDO No. 51 - Pasay City.xls",
    "RDO No. 52 - Paranaque City.xls",
    "RDO No. 43 - Pasig City.xls",
]

# Classification codes that indicate condos
CONDO_CODES = {"RC", "CC", "PS", "PH", "PC"}

# Keywords in street/name columns that indicate condos (more restrictive for row detection)
CONDO_KEYWORDS_STRICT = re.compile(
    r"\bCONDO\b|\bCONDOMINIUM\b|\bTOWER\b|\bBUILDING\b|\bBLDG\b|\bPENTHOUSE\b|"
    r"\bRESIDENCES\b|\bPARKING\b|\bFAR\s*[-=]|\bPER\s+SQ|\bFLOOR\s+AREA\b|"
    r"\bUNDER\s+CONSTRUCTION\b|\bTOWNHOUSE\b|\bSUITES?\b|\bMANSION\b|"
    r"\bPLAZA\b|\bGARDEN\b|\bVILLAS?\b|\bPALACE\b",
    re.IGNORECASE,
)


def cell_value_str(val):
    if val is None or val == "":
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def read_xls_sheet(filepath, sheet_index=1):
    wb = xlrd.open_workbook(str(filepath), formatting_info=True)
    if wb.nsheets <= sheet_index:
        sheet_index = 0
    sh = wb.sheet_by_index(sheet_index)
    rows = []
    for r in range(sh.nrows):
        row = []
        for c in range(sh.ncols):
            cell = sh.cell(r, c)
            row.append(cell_value_str(cell.value))
        rows.append(row)
    merged = list(sh.merged_cells)
    return rows, merged, sh.name, wb.nsheets, [wb.sheet_by_index(i).name for i in range(wb.nsheets)]


def read_xlsx_sheet(filepath, sheet_index=1):
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    sheet_names = wb.sheetnames
    if len(sheet_names) <= sheet_index:
        sheet_index = 0
    sh = wb[sheet_names[sheet_index]]
    rows = []
    for row in sh.iter_rows(min_row=1, max_row=sh.max_row, max_col=sh.max_column):
        rows.append([cell_value_str(cell.value) for cell in row])
    merged = []
    for mc in sh.merged_cells.ranges:
        merged.append((mc.min_row - 1, mc.max_row, mc.min_col - 1, mc.max_col))
    return rows, merged, sh.title, len(sheet_names), sheet_names


def get_merge_map(merged_cells, rows):
    """Build a map: (row, col) -> value from top-left of merge."""
    merge_map = {}
    for rlo, rhi, clo, chi in merged_cells:
        val = ""
        if rlo < len(rows) and clo < len(rows[rlo]):
            val = rows[rlo][clo]
        for r in range(rlo, rhi):
            for c in range(clo, chi):
                if r != rlo or c != clo:
                    merge_map[(r, c)] = val
    return merge_map


def get_effective_row(row_idx, rows, merge_map):
    """Get the effective values for a row, filling in merged cell values."""
    if row_idx >= len(rows):
        return []
    row = list(rows[row_idx])
    for c in range(len(row)):
        if not row[c] and (row_idx, c) in merge_map:
            row[c] = merge_map[(row_idx, c)]
    return row


def is_condo_section_header(text):
    """Check if text indicates start of a condo section."""
    upper = text.upper()
    return "LIST OF CONDOMINIUM" in upper or "LIST OF TOWNHOUSE" in upper or "CONDOMINIUM" in upper and "LIST" in upper


def format_row_compact(row, row_idx):
    """Format row showing all non-empty cells with column indices."""
    parts = []
    for i, val in enumerate(row):
        if val:
            parts.append(f"C{i}={val}")
    if not parts:
        return f"  R{row_idx}: (empty)"
    return f"  R{row_idx}: " + " | ".join(parts)


def find_condo_entries(rows, merged_cells, merge_map):
    """Find all condo entries, returning structured data."""
    entries = []
    current_zone = None
    in_condo_section = False

    for i, raw_row in enumerate(rows):
        eff_row = get_effective_row(i, rows, merge_map)
        row_text = " ".join(eff_row).upper()

        # Detect zone/barangay markers
        if any(kw in row_text for kw in ["ZONE ", "BARANGAY", "BRGY"]):
            for cell in eff_row:
                if cell and ("ZONE" in cell.upper() or "BARANGAY" in cell.upper() or "BRGY" in cell.upper()):
                    current_zone = cell.strip()
                    break

        # Detect condo section headers
        if is_condo_section_header(row_text):
            in_condo_section = True
            entries.append({
                'type': 'section_header',
                'row': i,
                'text': row_text.strip(),
                'zone': current_zone,
            })
            continue

        # Detect "(List of Streets)" headers - marks end of condo section for that zone page
        if "LIST OF STREETS" in row_text:
            in_condo_section = False
            continue

        # Check for condo classification codes
        has_condo_code = False
        code_found = None
        for cell in eff_row:
            val = cell.strip().upper()
            if val in CONDO_CODES:
                has_condo_code = True
                code_found = val
                break

        # Check for condo keywords
        has_keyword = bool(CONDO_KEYWORDS_STRICT.search(row_text))

        if has_condo_code or (in_condo_section and any(c for c in eff_row)):
            entries.append({
                'type': 'condo_row',
                'row': i,
                'raw': list(raw_row),
                'effective': list(eff_row),
                'code': code_found,
                'zone': current_zone,
                'in_section': in_condo_section,
                'keyword_match': has_keyword,
            })
        elif has_keyword and not in_condo_section:
            # Keyword match outside a marked condo section
            entries.append({
                'type': 'keyword_row',
                'row': i,
                'raw': list(raw_row),
                'effective': list(eff_row),
                'code': code_found,
                'zone': current_zone,
                'in_section': False,
            })

    return entries


def analyze_condo_structure(rows, merged_cells):
    """Analyze the overall condo table structure in a sheet."""
    # Find condo-related merged cells
    condo_merges = []
    for mc in merged_cells:
        rlo, rhi, clo, chi = mc
        if rlo < len(rows) and clo < len(rows[rlo]):
            val = rows[rlo][clo].upper()
            if any(kw in val for kw in ["CONDO", "TOWER", "BUILDING", "BLDG", "MANSION",
                                          "RESIDENCES", "SUITES", "PLAZA", "PLACE", "GARDEN",
                                          "PARKING", "UNIT", "PENTHOUSE"]):
                condo_merges.append((mc, rows[rlo][clo]))

    return condo_merges


def process_workbook(filepath, out):
    ext = filepath.suffix.lower()
    if ext == ".xls":
        rows, merged, sheet_name, nsheets, sheet_names = read_xls_sheet(filepath)
    elif ext == ".xlsx":
        rows, merged, sheet_name, nsheets, sheet_names = read_xlsx_sheet(filepath)
    else:
        out.write(f"  SKIP: Unknown extension {ext}\n")
        return

    num_rows = len(rows)
    num_cols = max(len(r) for r in rows) if rows else 0

    out.write(f"  Sheets: {nsheets} -> {sheet_names}\n")
    out.write(f"  Using sheet[1]: '{sheet_name}', {num_rows} rows x {num_cols} cols\n")
    out.write(f"  Total merged cell ranges: {len(merged)}\n\n")

    merge_map = get_merge_map(merged, rows)

    # ---- HEADER ANALYSIS ----
    out.write("  === COLUMN HEADER ANALYSIS ===\n")
    for i, row in enumerate(rows[:30]):
        row_text = " ".join(row).upper()
        if any(kw in row_text for kw in ["STREET NAME", "ZONE VALUE", "CLASS", "BARANGAY", "VICINITY",
                                          "SUBDIVISION", "MARKET VALUE", "CONDOMINIUM"]):
            out.write(f"  Header candidate R{i}:\n")
            for c, val in enumerate(row):
                if val:
                    out.write(f"    C{c}: {val}\n")
            # Also show effective values (with merges)
            eff = get_effective_row(i, rows, merge_map)
            if eff != row:
                out.write(f"  Effective (with merges):\n")
                for c, val in enumerate(eff):
                    if val:
                        out.write(f"    C{c}: {val}\n")
    out.write("\n")

    # ---- CONDO MERGED CELLS ----
    condo_merges = analyze_condo_structure(rows, merged)
    if condo_merges:
        out.write(f"  === CONDO-RELATED MERGED CELLS ({len(condo_merges)}) ===\n")
        for mc, val in condo_merges:
            rlo, rhi, clo, chi = mc
            out.write(f"    R{rlo}:{rhi} C{clo}:{chi} ({rhi-rlo}r x {chi-clo}c) = '{val}'\n")
        out.write("\n")

    # ---- ALL CONDO ENTRIES ----
    entries = find_condo_entries(rows, merged, merge_map)
    out.write(f"  === ALL CONDO ENTRIES ({len(entries)}) ===\n\n")

    # Group entries into condo buildings
    current_building = None
    building_entries = []

    for entry in entries:
        if entry['type'] == 'section_header':
            out.write(f"  ---- SECTION: {entry['text']} (R{entry['row']}, zone={entry['zone']}) ----\n\n")
            continue

        row_idx = entry['row']
        raw = entry['raw']
        eff = entry['effective']
        code = entry.get('code', '')

        # Print the row with full detail
        out.write(f"  R{row_idx} [code={code or '-'}, in_section={entry.get('in_section', '?')}]:\n")
        out.write(f"    RAW:  ")
        for c, val in enumerate(raw):
            if val:
                out.write(f"C{c}='{val}' ")
        out.write(f"\n")

        # Only print effective if different
        if eff != raw:
            out.write(f"    EFF:  ")
            for c, val in enumerate(eff):
                if val:
                    out.write(f"C{c}='{val}' ")
            out.write(f"\n")

        # Show which merged cells cover this row
        covering_merges = []
        for mc in merged:
            rlo, rhi, clo, chi = mc
            if rlo <= row_idx < rhi:
                if rlo < len(rows) and clo < len(rows[rlo]):
                    mval = rows[rlo][clo]
                else:
                    mval = ""
                covering_merges.append((mc, mval))
        if covering_merges:
            out.write(f"    MERGES: ")
            for mc, mval in covering_merges:
                rlo, rhi, clo, chi = mc
                out.write(f"[R{rlo}:{rhi} C{clo}:{chi}='{mval[:50]}'] ")
            out.write(f"\n")
        out.write(f"\n")

    # ---- PATTERN DETECTION: Look for groups of 3 rows per condo ----
    out.write(f"  === CONDO BLOCK PATTERN ANALYSIS ===\n")

    # Find all condo section headers and the rows after them
    for i, row in enumerate(rows):
        row_text = " ".join(row).upper()
        if "LIST OF CONDOMINIUM" in row_text or "LIST OF TOWNHOUSE" in row_text:
            out.write(f"\n  Condo section starts at R{i}: {row_text.strip()}\n")
            # Print the next 50 rows or until next section
            for j in range(i + 1, min(i + 60, len(rows))):
                eff = get_effective_row(j, rows, merge_map)
                eff_text = " ".join(eff).strip()
                if not eff_text:
                    out.write(f"    R{j}: (empty)\n")
                    continue
                # Check if we hit next page header
                if "STREET NAME" in eff_text.upper() or "REVENUE REGION" in eff_text.upper():
                    out.write(f"    R{j}: [SECTION BREAK] {eff_text[:80]}\n")
                    break
                out.write(f"    R{j}: ")
                for c, val in enumerate(eff):
                    if val:
                        out.write(f"C{c}='{val}' ")
                out.write(f"\n")

    out.write(f"\n  === RAW DUMP: ALL ROWS WITH RC/CC/PS/PH/PC CODES ===\n")
    for i, row in enumerate(rows):
        eff = get_effective_row(i, rows, merge_map)
        for c, val in enumerate(eff):
            if val.strip().upper() in CONDO_CODES:
                out.write(f"\n  R{i} (code={val.strip().upper()} in C{c}):\n")
                # Show 3 rows before and after with full cell detail
                for j in range(max(0, i - 3), min(len(rows), i + 4)):
                    e = get_effective_row(j, rows, merge_map)
                    marker = " >>>" if j == i else "    "
                    out.write(f"  {marker} R{j}: ")
                    for cc, vv in enumerate(e):
                        if vv:
                            out.write(f"C{cc}='{vv}' ")
                    out.write(f"\n")
                break  # only print once per row

    return entries


def main():
    outpath = OUT_DIR / "condo_extraction_results.txt"
    with open(outpath, "w") as out:
        out.write("=" * 120 + "\n")
        out.write("BIR ZONAL VALUE WORKBOOK - COMPREHENSIVE CONDO DATA EXTRACTION\n")
        out.write(f"Generated: 2026-03-03\n")
        out.write("=" * 120 + "\n\n")

        for filename in TARGET_FILES:
            filepath = BASE_DIR / filename
            if not filepath.exists():
                out.write(f"\nERROR: File not found: {filepath}\n")
                continue

            out.write(f"\n{'#' * 120}\n")
            out.write(f"# FILE: {filename}\n")
            out.write(f"{'#' * 120}\n\n")

            try:
                process_workbook(filepath, out)
            except Exception as e:
                out.write(f"  ERROR: {e}\n")
                import traceback
                traceback.print_exc(file=out)

            out.write(f"\n")

        out.write(f"\n{'=' * 120}\n")
        out.write("EXTRACTION COMPLETE\n")
        out.write(f"{'=' * 120}\n")

    print(f"Output written to: {outpath}")
    print(f"File size: {outpath.stat().st_size / 1024:.1f} KB")


if __name__ == "__main__":
    main()
